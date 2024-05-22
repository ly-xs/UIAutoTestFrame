from openpyxl import load_workbook
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from common.logger import Logger

logger = Logger(name="GetExcel").get_logger()


class GetExcel:
	"""读取excel文件数据"""

	def __init__(self, filepath, sheet_id=0, headlines=1):
		self.path = filepath
		self.sheet_id = sheet_id
		self.headlines = headlines

	def get_excel(self):
		"""
		加载json文件数据
		:return: 返回数据
		"""
		try:
			# 打开 path 对应的 excel文档
			wb = load_workbook(self.path)
			# 选择 sheet
			ws = wb.worksheets[self.sheet_id]
			data = []
			# 遍历 xlsx
			for row in ws.iter_rows(self.headlines):
				row = [str(cell.value) if not None else '' for cell in row]
				data.append(row)
			return data
		except Exception as msg:
			logger.error(f"excel文件获取数据异常{msg}")


# import xlrd
#
#
# class ReadExcel:
#     """读取excel文件数据"""
#
#     def __init__(self, fileName, SheetName="Sheet1"):
#         self.data = xlrd.open_workbook(fileName)
#         self.table = self.data.sheet_by_name(SheetName)
#
#         # 获取总行数、总列数
#         self.nrows = self.table.nrows
#         self.ncols = self.table.ncols
#
#     def read_data(self):
#         if self.nrows > 1:
#             # 获取第一行的内容，列表格式
#             keys = self.table.row_values(0)
#             listApiData = []
#             # 获取每一行的内容，列表格式
#             for col in range(1, self.nrows):
#                 values = self.table.row_values(col)
#                 # keys，values组合转换为字典
#                 api_dict = dict(zip(keys, values))
#                 listApiData.append(api_dict)
#             return listApiData
#         else:
#             print("表格是空数据!")
#             return None
