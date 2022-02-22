from openpyxl import load_workbook


def xstr(s):
	if s is None:
		return ''
	else:
		return str(s)


def get_data_from_excel(path, sheet_id=0, headlines=1):
	# 打开 path 对应的 excel文档
	wb = load_workbook(path)
	# 选择 sheet
	ws = wb.worksheets[sheet_id]
	content = []
	# 遍历 xlsx
	for row in ws.iter_rows(headlines):
		row = [xstr(cell.value) for cell in row]
		content.append(row)

	return content

# test
# ctx = get_data_from_excel(r'..\testdata\search_goods.xlsx',0,2)
# print(ctx)
